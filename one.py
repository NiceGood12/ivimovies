import requests
from bs4 import BeautifulSoup
import openpyxl

url = 'https://www.ivi.ru/animation/militants'

response = requests.get(url)

if response.status_code == 200:
    soup = BeautifulSoup(response.content, 'html.parser')

    movie_elements = soup.find_all('span', class_='nbl-slimPosterBlock__titleText')
    rating_elements = soup.find_all('div', class_='nbl-ratingCompact__value')
    info_elements = soup.find_all('div', class_='nbl-poster__propertiesInfo')

    movies = []

    for idx, element in enumerate(movie_elements):
        movie_title = element.text.strip()
        rating_element = rating_elements[idx]
        rating_integer = rating_element.find('div', class_='nbl-ratingCompact__valueInteger').text.strip()
        rating_fraction = rating_element.find('div', class_='nbl-ratingCompact__valueFraction').text.strip()

        info_element = info_elements[idx]
        info_rows = info_element.find_all('div', class_='nbl-poster__propertiesRow')
        year = info_rows[0].text.strip()
        country = info_rows[1].text.strip()

        movie_data = {
            'Title': movie_title,
            'Rating': f'{rating_integer}.{rating_fraction}',
            'Year': year,
            'Country': country
        }

        movies.append(movie_data)

    # Создание файла Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # Запись заголовков столбцов
    ws.cell(row=1, column=1, value='Title')
    ws.cell(row=1, column=2, value='Rating')
    ws.cell(row=1, column=3, value='Year')
    ws.cell(row=1, column=4, value='Country')

    # Запись данных фильмов
    for idx, movie in enumerate(movies, start=2):
        ws.cell(row=idx, column=1, value=movie['Title'])
        ws.cell(row=idx, column=2, value=movie['Rating'])
        ws.cell(row=idx, column=3, value=movie['Year'])
        ws.cell(row=idx, column=4, value=movie['Country'])

    # Сохранение файла Excel
    wb.save('movies.xlsx')
    print('Данные успешно сохранены в файле movies.xlsx.')
else:
    print('Не удалось получить доступ к странице.')
